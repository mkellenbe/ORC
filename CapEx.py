""" 
Calculate the Capital Expenditures using LandBOSSE, Wind Turbine Design Cost and Scaling Models 

The class CAPEX() includes functions that change parameters of the input to the
models depending on the location, rated power, rotor diameter and hub height,
after these changes the class also runs the models and reads out the final costs
"""

# Import the libraries used in the CAPEX() class
import os
from datetime import date
import math
import pandas as pd
# World bank database access python library
import wbgapi as wb
# Python library for modification and reading of Excel sheets
import openpyxl
# Python library for easy currency conversion
from currency_converter import CurrencyConverter

class CAPEX():

    def __init__(self, D_rotor_array, Power_rated_array, hub_height_array, ISO):
        """
        First we initialize the wind turbine parameters

        Args:
            D_rotor_array: User input rotor diameter
            Power_rated_array: User input rated power
            hub_height_array: User input hub height
            ISO: Alpha-3 code of the country in which the wind turbine is located in

        Returns:
            self: stores the attributes of the wind turbine within the class
        """
        self.D_rotor_array = D_rotor_array
        self.Power_rated_array = Power_rated_array
        self.hub_height_array = hub_height_array 
        self.ISO = ISO
    
    def changeInputLandBOSSE(self):
        """
        The function changes the data in the input file (Excel) for the LandBOSSE
        model according to the current wind turbine information input

        Args:
            self: attributes of the wind turbine

        Returns:
            modified Excel sheet that is later used as input to LandBOSSE
        """
        # Access the directory in which the project_list.xlsx file is 
        # (CHANGE FOR NEW USER DEPENDING ON LOCATION OF LANDBOSSE)
        # os.chdir(r"C:\Users\maxik\Desktop\ETH\Bachelorarbeit\Code\LCOE_model\landbosse\input")
        # Get access to the Excel sheet
        wb_append = openpyxl.load_workbook("landbosse\input\project_list.xlsx")
        sheet = wb_append.active
        # Change the variable values of the wind turbine
        sheet["D2"] = int(self.Power_rated_array/1000)
        sheet["E2"] = int(self.hub_height_array)
        sheet["F2"] = int(self.D_rotor_array)
        # Save the changes
        wb_append.save('landbosse\input\project_list.xlsx')

    def changeWageLandBOSSE(self):
        """
        The function changes the wages in the project data according to the location and occupation code
        in order to make the LandBOSSE model more region-specific

        A database for hourly rates all over the world gets used where possible,
        for all other cases a comparison to the 2019 USD hourly average wage is used
    
        Args:
            ISO: Alpha-3 code of the country in which the wind turbine is located in 
        
        Returns:
            modified Excel sheet that is later used as input to LandBOSSE
        """
        # Access the directory in which the project_list.xlsx file is 
        # (CHANGE FOR NEW USER DEPENDING ON LOCATION OF LANDBOSSE)
        # os.chdir(r"C:\Users\maxik\Desktop\ETH\Bachelorarbeit\Code\LCOE_model\landbosse\input\project_data")
        # Get access to the Excel sheet
        append = openpyxl.load_workbook("landbosse\input\project_data\project_test.xlsx")
        sheet = append["crew_price"]

        # Create a dataframe for the hourly salary in the world according to their codification (CHANGE FOR NEW USER DEPENDING ON LOCATION)
        df_original = pd.read_csv("databases\oww3.csv")
        wageList = df_original["isco88"]
        wageList1 = []
        # Filter out the / from the dataset, otherwise will not work
        for pay in wageList:
            wageList1.append(str(pay).split(sep="/"))
        serie = pd.Series(wageList1)
        df_original["isco88"] = serie
        # Only view data from the dataset, that is relevant to the location 
        df_original = df_original[df_original["country_code"] == self]
        # Limit the time of the data to only newer than 1995
        df_original = df_original[df_original["y0"]>1995]

        # Not all countries have data on all of the worker's hourly salaries, so for these cases a wage factor will be used to transform the costs uniformly
        # (CHANGE FOR NEW USER DEPENDING ON LOCATION OF THE DATASET)
        dat = pd.read_csv("databases\wages.csv")
        dat.sort_values(by="time",ascending=False)
        # We only want the average hourly wage, we assume that the price difference are proportional in each country as in the US
        dat = dat[dat["classif1"]=="OCU_SKILL_TOTAL"]
        dat1 = dat
        # There is no data in USD for Poland thus we make a tentative exception case
        if (self == 'POL'):
            dat = dat[dat["classif2"] == "CUR_TYPE_USD"]
            dat = dat[dat["ref_area"] == 'USA']
            US_wage = dat.iloc[0]['obs_value']
            # Data for Poland's average hourly wage gets converted into USD for ease of comparison
            dat1 = dat1[dat1["classif2"] == "CUR_TYPE_LCU"]
            dat1 = dat1[dat1["ref_area"] == self]
            ISO_wage_LCU = dat1.iloc[0]['obs_value']
            year = dat1.iloc[0]['time']
            c = CurrencyConverter()
            ISO_wage = c.convert(ISO_wage_LCU, 'PLN', 'USD', date=date(year, 6, 30))
        else:
            # For most other countries data in USD is available, so no currency exchange needs to be done
            dat = dat[dat["classif2"]=="CUR_TYPE_USD"]
            dat = dat[dat["ref_area"]=='USA']
            US_wage = dat.iloc[0]['obs_value']
            dat1 = dat1[dat1["ref_area"]==self]
            ISO_wage = dat1.iloc[0]['obs_value']
        # Finally a wage factor that describes the fraction of hourly pay at our specified location in comparison to the US is created
        wage_factor = ISO_wage/US_wage

        # Below all of the occupation's hourly wages through their respective ISCO88 codes are changed within the project_test.xlsx 
        # If data on a specific wage is missing we apply the wage_factor to get an estimate of the hourly wage

        # The available value signifies how many of the values in the Excel sheet get replaced through the database, NOT through the wage_factor
        available = 0

        # Only the first occupation will be described, as the rest are the same, just with a different ISCO88 value

        # Crane Operator ISIC:8333/9333
        try:
            df = df_original.copy()
            # The database is searched for the ISCO88 code
            mask = ['9333' in l for l in df["isco88"]]
            mask
            # Database now only consists of the wanted occupational hourly wages
            df=df[mask]
            # Sort the database by date to get most recent value
            df.sort_values(by="y0",ascending=False, inplace=True)
            # Get the wage that is the most recent
            wage = df.tail(1).iloc[0]['hw3wl_us']
            # Get the year at which the wage was inspected 
            year = int(df.tail(1).iloc[0]['y0'])
            # Adjust the value of the hourly wage in accordance to the year at which the value was inspected 
            # Write the value of the hourly wage into the Excel sheet
            sheet["B2"] = CAPEX.currentValue(wage,self,year)
            # The available data is added, as there now is one data point present
            available = available + 1
        except:
           sheet["B2"] = sheet["B2"].value*wage_factor

        # Oiler ISIC:8161
        try:
            df = df_original.copy()
            mask = ['8161' in l for l in df["isco88"]]
            mask
            df=df[mask]
            df.sort_values(by="y0",ascending=False)
            wage = df.tail(1).iloc[0]['hw3wl_us']
            year = int(df.tail(1).iloc[0]['y0'])
            sheet["B3"] = CAPEX.currentValue(wage,self,year)
            available = available + 1
        except:
            sheet["B3"] = sheet["B3"].value*wage_factor

        # Rigger ISIC:7215
        try:
            df = df_original.copy()
            mask = ['7215' in l for l in df["isco88"]]
            mask
            df=df[mask]
            df.sort_values(by="y0",ascending=False)
            wage = df.tail(1).iloc[0]['hw3wl_us']
            year = int(df.tail(1).iloc[0]['y0'])
            sheet["B4"] = CAPEX.currentValue(wage,self,year)
            available = available + 1
        except:
            sheet["B4"] = sheet["B4"].value*wage_factor

        # Truck driver ISIC:8324
        try:
            df = df_original.copy()
            mask = ['8324' in l for l in df["isco88"]]
            mask
            df=df[mask]
            df.sort_values(by="y0",ascending=False)
            wage = df.tail(1).iloc[0]['hw3wl_us']
            year = int(df.tail(1).iloc[0]['y0'])
            sheet["B5"] = CAPEX.currentValue(wage,self,year)
            available = available + 1
        except:
            sheet["B5"] = sheet["B5"].value*wage_factor

        # Iron worker ISIC:7222
        try:
            df = df_original.copy()
            mask = ['7222' in l for l in df["isco88"]]
            mask
            df=df[mask]
            df.sort_values(by="y0",ascending=False)
            wage = df.tail(1).iloc[0]['hw3wl_us']
            year = int(df.tail(1).iloc[0]['y0'])
            sheet["B6"] = CAPEX.currentValue(wage,self,year)
            available = available + 1
        except:
            sheet["B6"] = sheet["B6"].value*wage_factor

        # Project manager ISIC:1223
        try:
            df = df_original.copy()
            mask = ['1223' in l for l in df["isco88"]]
            mask
            df=df[mask]
            df.sort_values(by="y0",ascending=False)
            wage = df.tail(1).iloc[0]['hw3wl_us']
            year = int(df.tail(1).iloc[0]['y0'])
            sheet["B7"] = CAPEX.currentValue(wage,self,year)
            available = available + 1
        except:
            sheet["B7"] = sheet["B7"].value*wage_factor

        # Site manager ISIC:4330
        try:
            df = df_original.copy()
            mask = ['4330' in l for l in df["isco88"]]
            mask
            df=df[mask]
            df.sort_values(by="y0",ascending=False)
            wage = df.tail(1).iloc[0]['hw3wl_us']
            year = int(df.tail(1).iloc[0]['y0'])
            sheet["B8"] = CAPEX.currentValue(wage,self,year)
            available = available + 1
        except:
            sheet["B8"] = sheet["B8"].value*wage_factor

        # Construction manager ISIC:1313
        try:
            df = df_original.copy()
            mask = ['1313' in l for l in df["isco88"]]
            mask
            df=df[mask]
            df.sort_values(by="y0",ascending=False)
            wage = df.tail(1).iloc[0]['hw3wl_us']
            year = int(df.tail(1).iloc[0]['y0'])
            sheet["B9"] = CAPEX.currentValue(wage,self,year)
            available = available + 1
        except:
            sheet["B9"] = sheet["B9"].value*wage_factor

        # Project engineer ISIC:2142
        try:
            df = df_original.copy()
            mask = ['2142' in l for l in df["isco88"]]
            mask
            df=df[mask]
            df.sort_values(by="y0",ascending=False)
            wage = df.tail(1).iloc[0]['hw3wl_us']
            year = int(df.tail(1).iloc[0]['y0'])
            sheet["B10"] = CAPEX.currentValue(wage,self,year)
            available = available + 1
        except:
            sheet["B10"] = sheet["B10"].value*wage_factor

        # Safety manager ISIC:3112
        try:
            df = df_original.copy()
            mask = ['3112' in l for l in df["isco88"]]
            mask
            df=df[mask]
            df.sort_values(by="y0",ascending=False)
            wage = df.tail(1).iloc[0]['hw3wl_us']
            year = int(df.tail(1).iloc[0]['y0'])
            sheet["B11"] = CAPEX.currentValue(wage,self,year)
            available = available + 1
        except:
            sheet["B11"] = sheet["B11"].value*wage_factor

        # Logistics manager ISIC:1226
        try:
            df = df_original.copy()
            mask = ['1226' in l for l in df["isco88"]]
            mask
            df=df[mask]
            df.sort_values(by="y0",ascending=False)
            wage = df.tail(1).iloc[0]['hw3wl_us']
            year = int(df.tail(1).iloc[0]['y0'])
            sheet["B12"] = CAPEX.currentValue(wage,self,year)
            available = available + 1
        except:
            sheet["B12"] = sheet["B12"].value*wage_factor

        # Rigger foreman ISIC:3115
        try:
            df = df_original.copy()
            mask = ['3115' in l for l in df["isco88"]]
            mask
            df=df[mask]
            df.sort_values(by="y0",ascending=False)
            wage = df.tail(1).iloc[0]['hw3wl_us']
            year = int(df.tail(1).iloc[0]['y0'])
            sheet["B13"] = CAPEX.currentValue(wage,self,year)
            available = available + 1
        except:
            sheet["B13"] = sheet["B13"].value*wage_factor

        # Rigger ISIC:7215
        try:
            df = df_original.copy()
            mask = ['7215' in l for l in df["isco88"]]
            mask
            df=df[mask]
            df.sort_values(by="y0",ascending=False)
            wage = df.tail(1).iloc[0]['hw3wl_us']
            year = int(df.tail(1).iloc[0]['y0'])
            sheet["B14"] = CAPEX.currentValue(wage,self,year)
            available = available + 1
        except:
            sheet["B14"] = sheet["B14"].value*wage_factor

        # Operator ISIC:9333
        try:
            df = df_original.copy()
            mask = ['9333' in l for l in df["isco88"]]
            mask
            df=df[mask]
            df.sort_values(by="y0",ascending=False)
            wage = df.tail(1).iloc[0]['hw3wl_us']
            year = int(df.tail(1).iloc[0]['y0'])
            sheet["B15"] = CAPEX.currentValue(wage,self,year)
            available = available + 1
        except:
            sheet["B15"] = sheet["B15"].value*wage_factor

        # Oiler ISIC:8161
        try:
            df = df_original.copy()
            mask = ['8161' in l for l in df["isco88"]]
            mask
            df=df[mask]
            df.sort_values(by="y0",ascending=False)
            wage = df.tail(1).iloc[0]['hw3wl_us']
            year = int(df.tail(1).iloc[0]['y0'])
            sheet["B16"] = CAPEX.currentValue(wage,self,year)
            available = available + 1
        except:
            sheet["B16"] = sheet["B16"].value*wage_factor

        # Electrician ISIC:7137
        try:
            df = df_original.copy()
            mask = ['7137' in l for l in df["isco88"]]
            mask
            df=df[mask]
            df.sort_values(by="y0",ascending=False)
            wage = df.tail(1).iloc[0]['hw3wl_us']
            year = int(df.tail(1).iloc[0]['y0'])
            sheet["B17"] = CAPEX.currentValue(wage,self,year)
            available = available + 1
        except:
            sheet["B17"] = sheet["B17"].value*wage_factor

        # Tool room ISIC:7223
        try:
            df = df_original.copy()
            mask = ['7223' in l for l in df["isco88"]]
            mask
            df=df[mask]
            df.sort_values(by="y0",ascending=False)
            wage = df.tail(1).iloc[0]['hw3wl_us']
            year = int(df.tail(1).iloc[0]['y0'])
            sheet["B18"] = CAPEX.currentValue(wage,self,year)
            available = available + 1
        except:
            sheet["B18"] = sheet["B18"].value*wage_factor

        # QC\QA tech ISIC:3112
        try:
            df = df_original.copy()
            mask = ['3112' in l for l in df["isco88"]]
            mask
            df=df[mask]
            df.sort_values(by="y0",ascending=False)
            wage = df.tail(1).iloc[0]['hw3wl_us']
            year = int(df.tail(1).iloc[0]['y0'])
            sheet["B9"] = CAPEX.currentValue(wage,self,year)
            available = available + 1
        except:
            sheet["B19"] = sheet["B19"].value*wage_factor

        # Office admin ISIC:3439
        try:
            df = df_original.copy()
            mask = ['3439' in l for l in df["isco88"]]
            mask
            df=df[mask]
            df.sort_values(by="y0",ascending=False)
            wage = df.tail(1).iloc[0]['hw3wl_us']
            year = int(df.tail(1).iloc[0]['y0'])
            sheet["B20"] = CAPEX.currentValue(wage,self,year)
            available = available + 1
        except:
            sheet["B20"] = sheet["B20"].value*wage_factor

        
        # Print out the availability of data for the occupational hourly wage
        #print("Data on " + str(available) + " / 19 available")

        # Change the development cost with the wage factor
        sheet = append["development"]
        sheet["B3"] = sheet["B3"].value*wage_factor

        # Save the Excel sheet to represent region
        append.save("landbosse\input\project_data\project_test_changed.xlsx")

    def changeEquipPrice(self):
        """
        The function changes the price of the equipment needed to build the balance of system according to a to the PPP of the country,
        because prices of goods also vary across regions

        Args:
            ISO: Alpha-3 code of the country in which the wind turbine is located in

        Returns: 
            modified Excel sheet that is later used as input to LandBOSSE
        """
        # Navigate to the directory in which the input files for LandBOSSE are located
        # (CHANGE FOR NEW USER DEPENDING ON LOCATION OF LANDBOSSE)
        # os.chdir(r"C:\Users\maxik\Desktop\ETH\Bachelorarbeit\Code\LCOE_model\landbosse\input\project_data")
        # Get access to the Excel sheet
        append = openpyxl.load_workbook("landbosse\input\project_data\project_test_changed.xlsx")
        # Use the Price level ratio of PPP to adjust price in country in comparison to the USD
        sheet = append["equip_price"]
        currentPPP = wb.data.get(['PA.NUS.PPPC.RF'], self, 2019)
        currentPPP = currentPPP.get('value')
        # Iterate over all the Excel sheets with values 
        i = 2
        while i <= 15:
            sheet_name = "C"+ str(i)
            sheet[sheet_name] = sheet[sheet_name].value*currentPPP
            i = i + 1
        # Save the file
        append.save("landbosse\input\project_data\project_test_changed.xlsx")

    def perDiemLandBOSSE(self):
        """
        This function changes the per diem rates in different countries in the input excel sheet for LandBOSSE

        Args:
            ISO: Alpha-3 code of the country in which the wind turbine is located in
        
        Returns:
            changed Excel sheet according to the wind turbines location
        """
        # Navigate to the directory in which the input files for LandBOSSE are located
        # (CHANGE FOR NEW USER DEPENDING ON LOCATION OF LANDBOSSE)
        # os.chdir(r"C:\Users\maxik\Desktop\ETH\Bachelorarbeit\Code\LCOE_model\landbosse\input\project_data")
        append = openpyxl.load_workbook("landbosse\input\project_data\project_test_changed.xlsx")
        sheet = append["crew_price"]
        # The per diem rate gets also changed according to the location based on the European Commission's recommendations
        data = pd.read_csv("databases\per_diem.csv")
        data = data[data["ISO"] == self]
        perDiemRate = data.iloc[0]['Per diem rate']
        # And we convert it to the 2019 value in USD to have a common currency
        # The average 2017 EUR to $ exchange rate is used
        perDiemRate = 1.1301 * perDiemRate
        perDiemRate = CAPEX.currentValue(perDiemRate, self, 2017)
        # Iterate over all the sheets in the Excel files, where the per diem rate should be changed
        i = 2
        while i <= 21:
            sheet_name = "C"+ str(i)
            sheet[sheet_name] = perDiemRate
            i = i + 1
        # Save the Excel sheet
        append.save("landbosse\input\project_data\project_test_changed.xlsx")
 
    def runLandBOSSE(self):
        """
        The function is used to run the LandBOSSE model, which calculates the Balance of system costs and other metrics

        Args:
            self: attributes of the wind turbine
            input Excel sheets: project_test_changed.xlsx, project_list.xlsx

        Returns:
            output CSV sheets: landbosse-costs.csv            
        """
        CAPEX.changeInputLandBOSSE(self)
        # The directory has to be replaced when run elsewhere
        os.chdir("landbosse\LandBOSSE-2.5.0")
        # The main.py file runs the model 
        os.system(r"python main.py -i C:\Users\maxik\Desktop\ETH\Bachelorarbeit\Code\LCOE_model\landbosse\input -o C:\Users\maxik\Desktop\ETH\Bachelorarbeit\Code\LCOE_model\landbosse\output")

    def currentValue(wage, ISO, year):
        """
        This function converts a wage in a certain year to the 2019 inflation-adjusted wage, to better represent the actual cost.
        Using the world bank database for CPI (Consumer Price Index) we can calculate the current value of past wages.

        currentValue = pastValue * currentCPI / past CPI

        For this case the currentCPI is taken from the year 2019 to match the model's year of creation

        Args:
            wage: region-specific hourly wage
            ISO: Alpha-3 code of the country in which the wind turbine is located in
            year: year in which the wage was recorded

        Returns:
            currentWage: the region-specific hourly wage adjusted to 2019 USD
        """
        # Use the world database to get the CPI (Consumer Price Index) from the year 2019
        currentCPI = wb.data.get(['FP.CPI.TOTL'], ISO, 2019)
        # Use the world database to get the CPI (Consumer Price Index) from the year when the wage was recorded
        pastCPI = wb.data.get(['FP.CPI.TOTL'], ISO, year)
        # Usind the formula we can convert the value from the past wage to the current one
        currentWage = wage*currentCPI.get('value')/pastCPI.get('value')
        return(currentWage)

    def readCSVCosts():
        """
        This function reads out and sums the costs of the Landbosse model to get the overall cost of the Balance of system

        Returns:
            Total_cost_BOSSE: the total cost of all work and components needed for the balance of system
        """
        # First we specify the directory in which the output files are located at and then we open the newest directory 
        os.chdir("..\output")
        # List all subdirectories
        all_subdirs = [d for d in os.listdir('.') if os.path.isdir(d)]
        # Find the latest subdirectory based on modification time
        latest_subdir = max(all_subdirs, key=os.path.getmtime)
        # Construct the relative path to the CSV file
        relative_path_to_file = os.path.join(latest_subdir, "landbosse-costs.csv")
        # Read the CSV file using the relative path
        costs = pd.read_csv(relative_path_to_file)
        # After acquiring the data from the LandBOSSE model output we sum the costs to get the total cost for BOSSE 
        Total_cost_BOSSE = costs['Cost per turbine'].sum()
        return(Total_cost_BOSSE)

    def turbineCost(self):
        """
        turbine_costsse_2017.py

        Created by Arvydas Berzonskis  Goldwind 2017 based on turbine_costsse_2015.py 2015.
        Copyright (c) NREL. All rights reserved.

        This code was copied from the original made by Arvydas Berzonkis, as stated above.

        Here the cost of the wind turbine is calculated based on the rotor diameter, rated power and hub height.
        The code differs from the original, because not every calculated cost is needed, as these were already calculated through LandBOSSE.

        Args:
            self: attributes of the wind turbine

        Returns:
            turbineCost: total cost of the components of the wind turbine summed up
        """
        rotor_diameter = self.D_rotor_array
        machine_rating = self.Power_rated_array
        hub_height = self.hub_height_array

        # Calculate the blade mass and cost
        # Baseline mode
        self.blade_B_mass = 3 * 0.1452 * (rotor_diameter / 2)**2.9158  # all 3 blades
        # Advanced mode
        self.blade_A_mass = 0.4948 * (rotor_diameter / 2)**2.53  # all 3 blades
        # Blade material cost escalator
        self.BCE = 1
        # Labor cost escalator
        self.GDPE = 1
        # Costs
        self.blade_B_costs = 3 * ((0.4019 * (rotor_diameter / 2)**3 - 955.24) * self.BCE +
                                      2.7445 * (rotor_diameter / 2)**2.5025 * self.GDPE) / (1 - 0.28)
        self.blade_A_costs = 3 * ((0.4019 * (rotor_diameter / 2)**3 - 21051) * self.BCE +
                                      2.7445 * (rotor_diameter / 2)**2.5025 * self.GDPE) / (1 - 0.28)

        # calculate the Hub cost and weight
        self.hub_mass = 0.954 * (self.blade_B_mass / 3) + 5680.3
        self.hub_cost = self.hub_mass * 4.25

        # Pitch mechanisms and bearings
        self.pich_bearing_mass = 0.1295 * self.blade_B_mass + 491.31
        self.pitch_system_mass = self.pich_bearing_mass * 1.328 + 555
        # Total pitch costs
        self.pitch_system_cost = 2.28 * (0.2106 * rotor_diameter**2.6578)  # All 3 blades

        # Spinner, nose cone
        self.nose_cone_mass = 18.5 * rotor_diameter - 520.5
        self.nose_cone_cost = self.nose_cone_mass * 5.57

        # Low-speed shaft
        ''' Notes might not be used for direct drive turbine costs'''
        self.low_speed_shaft_mass = 0.0142 * rotor_diameter**2.888
        self.low_speed_shaft_cost = 0.01 * rotor_diameter**2.887

        # Main bearings
        self.bearing_mass = (rotor_diameter * 8 / 600 - 0.033) * 0.0092 * rotor_diameter**2.5
        self.bearing_cost = 2 * self.bearing_mass * 17.6

        # Mecahnical brake, high-speed coupling and associated components
        self.brake_and_coupling_cost = 1.9894 * machine_rating - 0.1141
        self.brake_and_coupling_mass = self.brake_and_coupling_cost / 10.

        # Direct drive Generator
        # self.generator_mass=661.25*self.low_speed_shaft_torque**0.606
        self.generator_cost = machine_rating * 219.33

        # Variable-speed electronics
        self.variablespeed_electronics = machine_rating * 79.

        # Yaw Drive and Bearing
        self.yaw_system_mass = 1.6 * (0.00098 * rotor_diameter**3.314)
        self.yaw_system_cost = 2 * (0.0339 * rotor_diameter**2.964)

        # Mainframe - Direct Drive
        self.mainframe_mass = 1.228 * rotor_diameter**1.953
        self.mainframe_cost = 627.28 * rotor_diameter**0.85

        # Platforms and railings
        self.platform_railing_mass = 0.125 * self.mainframe_mass
        self.platform_railing_cost = self.platform_railing_mass * 8.7

        # Electrical connections
        self.electrical_connection_cost = machine_rating * 40.

        # Hydraulic and Cooling Systems
        self.hydraulic_cooling_system_mass = 0.08 * machine_rating
        self.hydraulic_cooling_system_cost = machine_rating * 12

        # Nacelle Cover
        self.nacelle_cost = 11.537 * machine_rating + 3849.7
        self.nacelle_mass = self.nacelle_cost / 10.

        # Control, Safety Sytem, Condition Monitoring
        self.control_cost = 35000.0

        # Tower
        # Baseline model
        self.tower_B_mass = 0.3973 * (math.pi * (rotor_diameter / 2)**2) * hub_height - 1414
        self.tower_cost = self.tower_B_mass * 1.5

        turbineCost = (self.blade_B_costs + self.hub_cost + self.pitch_system_cost + 
                       self.nose_cone_cost + self.bearing_cost + self.brake_and_coupling_cost + 
                       self.generator_cost + self.variablespeed_electronics + self.yaw_system_cost + 
                       self.mainframe_cost + self.platform_railing_cost + self.electrical_connection_cost + 
                       self.hydraulic_cooling_system_cost + self.nacelle_cost + self.control_cost + self.tower_cost)

        return(CAPEX.currentValue(turbineCost, 'USA', 2006))

    def getCAPEX(self):
        """
        Here all the components of the class get combined to return the total CAPEX cost

        Args:
            self: attributes of the wind turbine
            ISO: Alpha-3 code of the country in which the wind turbine is located in

        Returns:
            CAPEXcost: total CAPEX cost
        """
        # Change values according to the region
        CAPEX.changeWageLandBOSSE(self.ISO)
        CAPEX.changeEquipPrice(self.ISO)
        CAPEX.perDiemLandBOSSE(self.ISO)
        # Run the landBOSSE model
        CAPEX.runLandBOSSE(self)
        # Add LandBOSSE and turbine costs together
        CAPEXcost = self.turbineCost() + CAPEX.readCSVCosts()
        return(CAPEXcost)
    
    def getCAPEXadjusted(self):
        """
        Here all the components of the class get combined to return the total CAPEX cost 
        The turbine costs are adjusted with a factor 

        Args:
            self: attributes of the wind turbine
            ISO: Alpha-3 code of the country in which the wind turbine is located in

        Returns:
            CAPEXcost: total CAPEX cost
        """
        # Change values according to the region
        CAPEX.changeWageLandBOSSE(self.ISO)
        CAPEX.changeEquipPrice(self.ISO)
        CAPEX.perDiemLandBOSSE(self.ISO)
        # Run the landBOSSE model
        CAPEX.runLandBOSSE(self)
        # Add LandBOSSE and turbine costs together, turbine costs adjusted with a factor 
        CAPEXcost = (40.01/92.3) * self.turbineCost() + CAPEX.readCSVCosts()
        return(CAPEXcost)
        

        




        




    
